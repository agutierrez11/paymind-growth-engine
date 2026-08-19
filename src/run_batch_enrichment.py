"""
Batch Lead Enrichment Worker for PayMind Growth Engine
Processes:
1. Hoteles_Boutique_256 (Sweet Spot: 10-60 rooms, Independent & Boutique)
2. Gasolineras_433 (Sweet Spot: 2-15 stations, Regional Franchises)
Integrates with SmartLeadRouter (Apollo, Hunter, Snov, SQLite Cache)
"""

import os
import sys
import time
import shutil

# Ensure UTF-8 output on Windows console
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

import pandas as pd
import openpyxl
from smart_lead_router import SmartLeadRouter

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "Campana_PayMind_MultiSegmento_CPS.xlsx")

# Exclusion sets for large corporate chains
EXCLUDED_CHAINS = [
    "marriott", "hilton", "ihg", "posadas", "fiesta americana", "fiesta inn", 
    "hyatt", "accor", "melia", "riu", "barcelo", "palace resorts", "hard rock",
    "oxxo gas", "petro-7", "hidrosina", "gulf mexico", "mobil corporativo"
]

def is_excluded(name: str, domain: str = "") -> bool:
    target = f"{name} {domain}".lower()
    for ex in EXCLUDED_CHAINS:
        if ex in target:
            return True
    return False

def run_enrichment():
    print("=" * 60)
    print("🚀 INICIANDO ENRIQUECIMIENTO INTELIGENTE MULTI-API (PAYMIND)")
    print("=" * 60)
    
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ No se encontró el archivo: {EXCEL_PATH}")
        return

    router = SmartLeadRouter()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # -------------------------------------------------------------
    # 1. PROCESAR HOTELES BOUTIQUE (256 LEADS)
    # -------------------------------------------------------------
    if "Hoteles_Boutique_256" in wb.sheetnames:
        print("\n🏨 [1/2] Procesando Lote: Hoteles Boutique e Independientes...")
        sheet = wb["Hoteles_Boutique_256"]
        headers = [cell.value for cell in sheet[1]]
        
        # Mapear columnas o agregar columnas de enriquecimiento
        col_map = {h: i + 1 for i, h in enumerate(headers) if h}
        
        # Agregar columnas si no existen
        extra_cols = ["Contacto_Enriquecido", "Cargo_Decisor", "Email_Verificado", "Fuente_Enriquecimiento", "Status_Hunter_Snov"]
        for col_name in extra_cols:
            if col_name not in col_map:
                col_idx = len(headers) + 1
                headers.append(col_name)
                sheet.cell(row=1, column=col_idx, value=col_name)
                col_map[col_name] = col_idx

        # Iterar filas (desde la 2)
        total_rows = sheet.max_row
        print(f"Total de registros a evaluar: {total_rows - 1}")
        
        enriched_count = 0
        skipped_chains = 0
        
        for r in range(2, min(total_rows + 1, 60)): # Procesar lote prioritario
            hotel_name = str(sheet.cell(row=r, column=col_map.get("Hotel_Nombre", 1)).value or "").strip()
            email = str(sheet.cell(row=r, column=col_map.get("Correo_General", 2)).value or "").strip()
            domain = str(sheet.cell(row=r, column=col_map.get("Dominio", 3)).value or "").strip()
            
            if not hotel_name or hotel_name == "None":
                continue
                
            if is_excluded(hotel_name, domain):
                skipped_chains += 1
                sheet.cell(row=r, column=col_map["Fuente_Enriquecimiento"], value="Descartado (Cadena Grande)")
                continue

            # Buscar decisores (Gerente General, Operaciones, Dueño)
            lead_data = {
                "Compañia": hotel_name,
                "Email": email if "@" in email else f"info@{domain}" if domain else "",
                "Dominio": domain
            }
            
            # Smart Routing
            print(f"  [{r-1}] Evaluando: {hotel_name} ({domain or email})...", end=" ")
            res = router.enrich_lead_smart(lead_data)
            
            contacts = res.get("decision_makers", [])
            if contacts:
                best = contacts[0]
                sheet.cell(row=r, column=col_map["Contacto_Enriquecido"], value=best.get("name", ""))
                sheet.cell(row=r, column=col_map["Cargo_Decisor"], value=best.get("title", ""))
                sheet.cell(row=r, column=col_map["Email_Verificado"], value=best.get("email", "") or email)
                sheet.cell(row=r, column=col_map["Fuente_Enriquecimiento"], value="Apollo / Smart Match")
                sheet.cell(row=r, column=col_map["Status_Hunter_Snov"], value="Decision Maker Found")
                print(f"✅ Match: {best.get('name')} ({best.get('title')})")
                enriched_count += 1
            else:
                # Verificación de email existente con Hunter
                v_status = res.get("verified_status", "unverified")
                sheet.cell(row=r, column=col_map["Email_Verificado"], value=email)
                sheet.cell(row=r, column=col_map["Fuente_Enriquecimiento"], value="Hunter.io / Direct")
                sheet.cell(row=r, column=col_map["Status_Hunter_Snov"], value=v_status)
                print(f"ℹ️ Verificación: {v_status}")

            # Pequeña pausa para respetar rate limits de APIs
            time.sleep(0.4)

        print(f"\n✨ Hoteles procesados: {enriched_count} decisores encontrados | {skipped_chains} cadenas excluidas.")

    # -------------------------------------------------------------
    # 2. PROCESAR GASOLINERAS (433 ESTACIONES)
    # -------------------------------------------------------------
    if "Gasolineras_433" in wb.sheetnames:
        print("\n⛽ [2/2] Procesando Lote: Gasolineras y Grupos Regionales...")
        sheet_gas = wb["Gasolineras_433"]
        headers_gas = [cell.value for cell in sheet_gas[1]]
        col_map_gas = {h: i + 1 for i, h in enumerate(headers_gas) if h}
        
        extra_cols = ["Contacto_Enriquecido", "Cargo_Decisor", "Email_Verificado", "Fuente_Enriquecimiento", "Status_Hunter_Snov"]
        for col_name in extra_cols:
            if col_name not in col_map_gas:
                col_idx = len(headers_gas) + 1
                headers_gas.append(col_name)
                sheet_gas.cell(row=1, column=col_idx, value=col_name)
                col_map_gas[col_name] = col_idx

        total_gas = sheet_gas.max_row
        print(f"Total registros gasolineras: {total_gas - 1}")
        
        enriched_gas = 0
        for r in range(2, min(total_gas + 1, 60)): # Lote inicial
            estacion_name = str(sheet_gas.cell(row=r, column=col_map_gas.get("Nombre_Comercial", 1)).value or "").strip()
            estado = str(sheet_gas.cell(row=r, column=col_map_gas.get("Estado", 2)).value or "").strip()
            
            if not estacion_name or estacion_name == "None":
                continue
                
            if is_excluded(estacion_name):
                sheet_gas.cell(row=r, column=col_map_gas["Fuente_Enriquecimiento"], value="Descartado (Megagrupo)")
                continue

            print(f"  [{r-1}] Buscando grupo: {estacion_name} ({estado})...", end=" ")
            
            # Buscar en Apollo por nombre de grupo / estación
            apollo_res = router.search_apollo_decision_makers(
                estacion_name, 
                titles=["Director", "Gerente", "Dueño", "Finanzas", "Administrador", "Contador"]
            )
            contacts = apollo_res.get("contacts", [])
            if contacts:
                best = contacts[0]
                sheet_gas.cell(row=r, column=col_map_gas["Contacto_Enriquecido"], value=best.get("name", ""))
                sheet_gas.cell(row=r, column=col_map_gas["Cargo_Decisor"], value=best.get("title", ""))
                sheet_gas.cell(row=r, column=col_map_gas["Email_Verificado"], value=best.get("email", ""))
                sheet_gas.cell(row=r, column=col_map_gas["Fuente_Enriquecimiento"], value="Apollo / Smart Match")
                sheet_gas.cell(row=r, column=col_map_gas["Status_Hunter_Snov"], value="C-Level / Admin Found")
                print(f"✅ Match: {best.get('name')} ({best.get('title')})")
                enriched_gas += 1
            else:
                sheet_gas.cell(row=r, column=col_map_gas["Fuente_Enriquecimiento"], value="Direct / Regional")
                sheet_gas.cell(row=r, column=col_map_gas["Status_Hunter_Snov"], value="Pending Manual Call")
                print("ℹ️ Pendiente contacto directo")

            time.sleep(0.4)

        print(f"\n✨ Gasolineras procesadas: {enriched_gas} decisores encontrados.")

    # Guardar cambios en el Excel Master
    wb.save(EXCEL_PATH)
    print(f"\n💾 Archivo Master guardado con éxito en: {EXCEL_PATH}")

    # Copiar a Descargas y Escritorio
    destinations = [
        r"C:\Users\Antonio\Downloads\Campana_PayMind_MultiSegmento_CPS.xlsx",
        r"C:\Users\Antonio\OneDrive\Downloads\Campana_PayMind_MultiSegmento_CPS.xlsx",
        r"C:\Users\Antonio\Desktop\Campana_PayMind_MultiSegmento_CPS.xlsx"
    ]
    for d in destinations:
        try:
            if os.path.exists(os.path.dirname(d)):
                shutil.copy2(EXCEL_PATH, d)
                print(f"📋 Copia actualizada en: {d}")
        except Exception as e:
            pass

    print("\n🎉 ¡ENRIQUECIMIENTO COMPLETADO CON ÉXITO!")

if __name__ == "__main__":
    run_enrichment()
