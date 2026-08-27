"""
Generador Maestro de la Campaña PayMind Multi-Segmento (CPS & Venta Socrática)
Crea un archivo Excel profesional con hojas segmentadas para:
1. Hoteles Boutique e Independientes (256 leads limpios)
2. Gasolineras Regionales y Franquiciatarios (433 estaciones)
3. Escuelas y Colegios Privados K-12
4. Clínicas y Consultorios Médicos Especializados
5. Playbook Metodológico CPS & Venta Socrática
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"C:\Users\Antonio\.gemini\antigravity-ide\scratch\paymind-growth-engine"
excel_out = os.path.join(base_dir, "data", "Campana_PayMind_MultiSegmento_CPS.xlsx")

# 1. HOTELES BOUTIQUE (256 LEADS DEDUPLICADOS)
hoteles_csv = os.path.join(base_dir, "data", "hoteles_deduplicados.csv")
df_h = pd.read_csv(hoteles_csv, encoding='utf-8-sig')

df_h['Segmento'] = 'Hotel Independiente / Boutique'
df_h['Decisor_Objetivo'] = 'Director General / Administrador / Contralor'

df_h['Asunto_Paso1_Apertura'] = df_h['Hotel_Nombre'].apply(
    lambda x: f"{x} — Reservaciones caídas y comisiones en Front-Desk"
)
df_h['Cuerpo_Paso1_Apertura'] = df_h['Hotel_Nombre'].apply(
    lambda x: (
        f"Hola,\n\n"
        f"Analizando la operación de cobros en {x}, un dolor recurrente en hoteles boutique e independientes es la fricción en Front-Desk: "
        f"terminales bancarias que tardan en autorizar depósitos de garantía, comisiones elevadas en cobro de restaurante por separado "
        f"y el riesgo constante de contracargos en reservaciones telefónicas o por WhatsApp.\n\n"
        f"En PayMind ayudamos a hoteles independientes a resolver esto con:\n"
        f"1. SmartPOS Android 4G Portátil: Pre-autorizaciones rápidas de garantía y cobro unificado (Recepción + Restaurante/Bar) en un solo dispositivo inalámbrico.\n"
        f"2. Ligas de Pago Seguras por WhatsApp: Cobro de anticipos con autenticación 3D Secure 2.0 que transfiere la responsabilidad del fraude al banco emisor (Cero contracargos).\n"
        f"3. Meses Sin Intereses (MSI): Con todos los bancos para estancias largas.\n\n"
        f"¿Tendrás 10 minutos este jueves para ver cómo eliminar la fricción de cobro y reducir costos de adquirencia en {x}?\n\n"
        f"Saludos,\nAntonio Gutiérrez\nConsultor de Crecimiento & Pagos | PayMind"
    )
)

df_h['Asunto_Paso2_FollowUp_D4'] = df_h['Hotel_Nombre'].apply(
    lambda x: f"Re: {x} — Cobro de anticipos sin riesgo de contracargo"
)
df_h['Cuerpo_Paso2_FollowUp_D4'] = df_h['Hotel_Nombre'].apply(
    lambda x: (
        f"Hola,\n\n"
        f"Te escribo rápido en seguimiento a mi correo sobre {x}.\n\n"
        f"Cuando un huésped reserva directo por teléfono o WhatsApp, ¿cómo cobran hoy el anticipo sin exponerse a que desconozcan el cargo después? "
        f"Nuestra pasarela genera ligas tokenizadas con validación bancaria directa que blindan el cobro al 100%.\n\n"
        f"¿Te hace sentido revisar una comparativa de 10 minutos esta semana?\n\n"
        f"Saludos,\nAntonio Gutiérrez"
    )
)

df_h['Asunto_Paso3_CasoExito_D8'] = df_h['Hotel_Nombre'].apply(
    lambda x: f"Caso de Éxito: 35% menos tiempo en Check-In y cobro unificado"
)
df_h['Cuerpo_Paso3_CasoExito_D8'] = df_h['Hotel_Nombre'].apply(
    lambda x: (
        f"Hola,\n\n"
        f"Te comparto un dato puntual: hoteles de 15 a 60 habitaciones que migraron sus cobros de Front-Desk a PayMind SmartPOS redujeron "
        f"35% el tiempo de espera de los huéspedes en recepción y consolidaron los reportes de ventas de habitaciones y consumos en un solo dashboard en tiempo real.\n\n"
        f"¿Platicamos brevemente este viernes?\n\n"
        f"Saludos,\nAntonio Gutiérrez"
    )
)

df_h['Asunto_Paso4_Breakup_D12'] = df_h['Hotel_Nombre'].apply(
    lambda x: f"¿Cerramos el tema por ahora, {x}?"
)
df_h['Cuerpo_Paso4_Breakup_D12'] = df_h['Hotel_Nombre'].apply(
    lambda x: (
        f"Hola,\n\n"
        f"Asumo que por el momento tienen resuelta la parte de terminales y pasarelas de pago en {x}.\n\n"
        f"Dejo la puerta abierta: si más adelante buscan optimizar tasas de adquirencia, habilitar MSI o blindarse contra contracargos, "
        f"puedes contactarme directamente por este medio.\n\n"
        f"Mucho éxito en la temporada.\n\n"
        f"Antonio Gutiérrez"
    )
)

# 2. GASOLINERAS REGIONALES (433 ESTACIONES CON MATEMÁTICA FISCAL DEL 40%)
gas_xlsx = os.path.join(base_dir, "data", "Campana_PayMind_Gasolineras_400.xlsx")
if os.path.exists(gas_xlsx):
    df_g = pd.read_excel(gas_xlsx, sheet_name=0)
else:
    df_g = pd.read_csv(os.path.join(base_dir, "data", "Campana_PayMind_Gasolineras_400.csv"), encoding='latin1')

df_g.columns.values[0] = 'Compania'
df_g['Compania'] = df_g['Compania'].astype(str).str.strip()
df_g['Nombre'] = df_g['Nombre'].astype(str).str.strip()

df_g['Asunto_Paso1_Apertura'] = df_g['Compania'].apply(
    lambda x: f"{x} — Agilizar cobro en bombas sin cambiar tu sistema actual"
)
df_g['Cuerpo_Paso1_Apertura'] = df_g.apply(
    lambda r: (
        f"Hola {r['Nombre']},\n\n"
        f"Platicando con dueños de estaciones y cadenas regionales, un dolor recurrente en horas de alto tráfico es ver conductores que se desesperan por la demora en el cobro y se van a cargar a la gasolinera de enfrente.\n\n"
        f"En PayMind ayudamos a grupos gasolineros como {r['Compania']} a acelerar el cobro en bomba sin pedirte que cambies tu sistema de control ni tu proveedor actual:\n"
        f"1. Terminales inalámbricas por bomba: El despachador cobra en segundos sin alejarse del auto.\n"
        f"2. Cero captura manual: La bomba le manda el monto exacto a la terminal, evitando errores de dedo ante el SAT.\n"
        f"3. Unificación de vales y tarjetas: Recibes tarjetas bancarias y vales (Edenred, Sodexo, SiVale) en el mismo equipo con dinero disponible al día siguiente (T+1).\n\n"
        f"Nos adaptamos a tu infraestructura actual (o si lo prefieres, te conectamos con nuestro socio de telemetría).\n\n"
        f"¿Tendrás 10 minutos este jueves para revisar cómo agilizar el cobro en tus estaciones sin arriesgar la operación?\n\n"
        f"Saludos,\nAntonio Gutiérrez\nConsultor de Crecimiento & Pagos | PayMind"
    ), axis=1
)

df_g['Asunto_Paso2_FollowUp_D4'] = df_g['Compania'].apply(
    lambda x: f"Re: {x} — Despachadores pidiéndose terminales prestadas entre bombas"
)
df_g['Cuerpo_Paso2_FollowUp_D4'] = df_g.apply(
    lambda r: (
        f"Hola {r['Nombre']},\n\n"
        f"Te escribo brevemente en seguimiento.\n\n"
        f"Cuando hay 6 u 8 bombas y solo 2 o 3 terminales bancarias, los despachadores pierden tiempo caminando a pedírsela prestada al compañero. Eso genera filas y hace que el cobro tome minutos en lugar de segundos.\n\n"
        f"Con nuestras terminales inalámbricas ligeras, cada despachador cobra al instante sin caminar ni teclear montos a mano (lo que elimina discrepancias de facturación).\n\n"
        f"¿Te interesaría ver una demo técnica de 10 minutos esta semana?\n\n"
        f"Saludos,\nAntonio Gutiérrez"
    ), axis=1
)

df_g['Asunto_Paso3_CasoExito_D8'] = df_g['Compania'].apply(
    lambda x: f"Cierres de turno a las 2:00 PM sin frenar la atención en {x}"
)
df_g['Cuerpo_Paso3_CasoExito_D8'] = df_g.apply(
    lambda r: (
        f"Hola {r['Nombre']},\n\n"
        f"Un problema común en estaciones de servicio es el cambio de turno de las 2:00 PM: justo en hora pico de tráfico, la atención se alenta porque los despachadores se detienen a contar vouchers y hacer cortes manuales.\n\n"
        f"Grupos gasolineros que unificaron sus cobros con PayMind hacen cierres digitales en 2 minutos desde la terminal, además de recibir el dinero de vales y tarjetas al día siguiente (T+1) para comprar pipas.\n\n"
        f"¿Platicamos 10 minutos este viernes?\n\n"
        f"Saludos,\nAntonio Gutiérrez"
    ), axis=1
)

df_g['Asunto_Paso4_Breakup_D12'] = df_g['Compania'].apply(
    lambda x: f"{x} — Dejando esto listo para cuando revisen terminales"
)
df_g['Cuerpo_Paso4_Breakup_D12'] = df_g.apply(
    lambda r: (
        f"Hola {r['Nombre']},\n\n"
        f"Entiendo que por ahora tienen resuelta la parte de cobro con su banco adquirente actual.\n\n"
        f"Te dejo nuestro enlace por si más adelante buscan reducir comisiones bancarias o conectar terminales inalámbricas directo a sus bombas sin cambiar de sistema:\n"
        f"https://paymind-growth-engine.pages.dev/\n\n"
        f"Mucho éxito en la operación de {r['Compania']}.\n\n"
        f"Saludos,\nAntonio Gutiérrez"
    ), axis=1
)

# 3. ESCUELAS Y COLEGIOS PRIVADOS (+45k Universo)
df_escuelas = pd.DataFrame([
    {
        'Institucion_Nombre': 'Colegio Bilingüe San Jerónimo',
        'Correo': 'direccion.general@sanjeronimo.edu.mx',
        'Segmento': 'Educación K-12 / Privada',
        'Decisor_Objetivo': 'Director General / Administrador',
        'Asunto_Paso1_Apertura': 'Colegio San Jerónimo — Cobro de colegiaturas con MSI y ligas WhatsApp',
        'Cuerpo_Paso1_Apertura': (
            'Hola,\n\n'
            'A inicio de cada ciclo o mes, las escuelas privadas enfrentan filas en caja, terminales lentas y familias solicitando Meses Sin Intereses (MSI).\n\n'
            'En PayMind facilitamos el cobro de colegiaturas con SmartPOS en ventanilla y ligas de pago por WhatsApp con hasta 12 MSI en todos los bancos.\n\n'
            '¿Tendrás 10 minutos para revisar cómo agilizar la cobranza este ciclo?\n\n'
            'Saludos,\nAntonio Gutiérrez | PayMind'
        ),
        'Asunto_Paso2_FollowUp_D4': 'Re: Colegio San Jerónimo — Reducción de mora en colegiaturas',
        'Cuerpo_Paso2_FollowUp_D4': 'Hola,\n\nNuestras ligas automatizadas reducen la morosidad de pago en un 25% al permitir a los padres pagar desde su celular con tarjeta de crédito o débito.\n\n¿Platicamos 10 minutos esta semana?\n\nSaludos.',
        'Asunto_Paso3_CasoExito_D8': 'Caso Colegio Privado: 0 filas en caja a inicio de mes',
        'Cuerpo_Paso3_CasoExito_D8': 'Hola,\n\nColegios con más de 300 alumnos migraron sus cobros a PayMind eliminando el 100% de las filas de caja y recibiendo dispersión inmediata.\n\n¿Coordinamos una llamada breve?\n\nSaludos.',
        'Asunto_Paso4_Breakup_D12': '¿Dejamos el tema en pausa por ahora?',
        'Cuerpo_Paso4_Breakup_D12': 'Hola,\n\nEntiendo que están cubiertos en su pasarela de pagos actual. Quedo a su disposición para el siguiente ciclo escolar.\n\nMucho éxito.'
    }
])

# 4. CLÍNICAS Y CONSULTORIOS MÉDICOS (+60k Universo)
df_clinicas = pd.DataFrame([
    {
        'Clinica_Nombre': 'Centro Odontológico & Estética Integral',
        'Correo': 'administracion@dentalintegral.com',
        'Segmento': 'Salud / Clínicas Especializadas',
        'Decisor_Objetivo': 'Director Médico / Socio Administrador',
        'Asunto_Paso1_Apertura': 'Centro Odontológico — Cierre de tratamientos de alto ticket con 12 MSI',
        'Cuerpo_Paso1_Apertura': (
            'Hola,\n\n'
            'En tratamientos médicos y dentales de $5,000 a $30,000 MXN, el 60% de los pacientes decide iniciar si se les ofrecen Meses Sin Intereses (MSI) con su tarjeta.\n\n'
            'En PayMind equipamos a clínicas con SmartPOS 4G inalámbricas con 3, 6, 9 y 12 MSI en todos los bancos y comisiones preferenciales.\n\n'
            '¿Te hace sentido revisar una demo de 10 minutos?\n\n'
            'Saludos,\nAntonio Gutiérrez | PayMind'
        ),
        'Asunto_Paso2_FollowUp_D4': 'Re: Centro Odontológico — Aceptación de todos los bancos en tratamientos',
        'Cuerpo_Paso2_FollowUp_D4': 'Hola,\n\n¿Sabías que muchas terminales bancarias tradicionales restringen tarjetas de ciertos bancos en promociones a meses? PayMind acepta Visa, Mastercard, Amex y Carnet sin trabas.\n\n¿Tienes 10 minutos este jueves?\n\nSaludos.',
        'Asunto_Paso3_CasoExito_D8': 'Caso Clínica: +28% en tasa de cierre de tratamientos',
        'Cuerpo_Paso3_CasoExito_D8': 'Hola,\n\nClínicas de especialidades aumentaron su cierre de presupuestos un 28% al ofrecer MSI directo en el sillón de consulta con la SmartPOS inalámbrica.\n\n¿Platicamos brevemente?\n\nSaludos.',
        'Asunto_Paso4_Breakup_D12': '¿Cerramos el contacto por ahora?',
        'Cuerpo_Paso4_Breakup_D12': 'Hola,\n\nEntiendo que su esquema de cobro actual está cubierto. Quedo a la orden cuando busquen renovar terminales o habilitar MSI.\n\nSaludos.'
    }
])

# 5. PLAYBOOK METODOLÓGICO CPS Y VENTA SOCRÁTICA
df_cps = pd.DataFrame([
    {
        'Vertical': 'Gasolineras Regionales (2-15 Estaciones)',
        'Situacion_Actual': 'Cobran con terminales BBVA/Banorte; tienen 4 maquinitas de vales colgadas; capturan montos a mano.',
        'Problema_Raiz_CPS (First Principles)': 'El banco tradicional cobra 2% sobre los $24 pesos totales con 40% de impuestos, quedándose con el 43% de la ganancia neta. Riesgo de multas por Anexo 30 del SAT.',
        'Pregunta_Socratica_Apertura': '¿Cuánto margen neto le está quitando su banco actual al cobrar comisiones sobre el IEPS y el IVA que usted no gana?',
        'Solucion_PayMind': 'SmartPOS única en bomba con App2App al dispensario + Todos los vales (Edenred/Sodexo) + Liquidación T+1 para pipas.'
    },
    {
        'Vertical': 'Hoteles Boutique & Independientes',
        'Situacion_Actual': 'Terminal fija en recepción; cobro de anticipos por transferencia o teléfono con riesgo de fraude.',
        'Problema_Raiz_CPS (First Principles)': 'Fricción en Front-Desk, caídas de señal en restaurante y vulnerabilidad a contracargos en reservas directas sin 3DS.',
        'Pregunta_Socratica_Apertura': '¿Cuál es el porcentaje de reservaciones directas que sufren contracargo o fricción en el depósito de garantía?',
        'Solucion_PayMind': 'Ligas tokenizadas con 3DS2 + SmartPOS inalámbrica 4G para Recepción y Restaurante.'
    },
    {
        'Vertical': 'Escuelas y Colegios Privados',
        'Situacion_Actual': 'Ventanilla saturada a inicio de mes; padres piden diferir pagos a meses.',
        'Problema_Raiz_CPS (First Principles)': 'Falta de opciones digitales multicanal que incentiven el pronto pago sin absorber costo operativo de cobranza.',
        'Pregunta_Socratica_Apertura': '¿Cuántas horas de personal administrativo se consumen cada quincena atendiendo filas de cobro en caja?',
        'Solucion_PayMind': 'Cobro WhatsApp + Terminal ventanilla + MSI con todos los bancos emisores.'
    },
    {
        'Vertical': 'Clínicas y Consultorios Médicos',
        'Situacion_Actual': 'Pacientes posponen cirugías o tratamientos estéticos/dentales por falta de financiamiento.',
        'Problema_Raiz_CPS (First Principles)': 'Pérdida de conversión en presupuestos de alto ticket ($5k - $30k) por no ofrecer MSI en el punto de consulta.',
        'Pregunta_Socratica_Apertura': '¿Cuántos presupuestos de más de $10,000 pesos se quedan sin cerrar por no ofrecer 12 meses sin intereses?',
        'Solucion_PayMind': 'SmartPOS inalámbrica con 12 MSI universal y liquidación inmediata.'
    }
])

# GUARDAR MASTER WORKBOOK
with pd.ExcelWriter(excel_out, engine='openpyxl') as writer:
    df_h.to_excel(writer, sheet_name='Hoteles_Boutique_256', index=False)
    df_g.to_excel(writer, sheet_name='Gasolineras_433', index=False)
    df_escuelas.to_excel(writer, sheet_name='Escuelas_Colegios', index=False)
    df_clinicas.to_excel(writer, sheet_name='Clinicas_Medicas', index=False)
    df_cps.to_excel(writer, sheet_name='Playbook_CPS_Metodologia', index=False)

# ESTILIZAR CON OPENPYXL
wb = openpyxl.load_workbook(excel_out)

header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

for ws in wb.worksheets:
    ws.views.sheetView[0].showGridLines = True
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:15]:
            val_str = str(cell.value or '')
            max_len = max(max_len, min(len(val_str), 45))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

wb.save(excel_out)
print(f"Master Workbook generado exitosamente en: {excel_out}")
